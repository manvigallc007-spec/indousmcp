"""H-1B intelligence from the free DOL OFLC LCA disclosure data → Dost's knowledge base.

The U.S. Department of Labor publishes every H-1B Labor Condition Application (employer, job title,
worksite, offered wage) as public-domain disclosure files. H-1B is the visa most Indians-from-India
use to work in the USA (~70% of holders are Indian nationals) and the largest sponsors are largely
Indian IT firms — so these aggregates are core diaspora "professional & income" intelligence:
typical wages by occupation, the biggest sponsoring employers, and where the jobs are.

We stream the (large) official file row-by-row, keep only **certified H-1B** rows, compute aggregates,
and store them as knowledge documents. No PII is kept — only public, aggregated figures.

Operator-driven, like the caterbid import: set DOL_H1B_DISCLOSURE_URL (a dol.gov URL or a local path)
to the current fiscal-year file; blank = disabled. .csv streams in constant memory (stdlib); .xlsx
needs openpyxl. The annual file is huge — prefer one quarter, or convert to .csv if RAM is tight.
"""

from __future__ import annotations

import os
import statistics
import tempfile
from collections import Counter, defaultdict
from typing import Any, Iterator

import httpx

from .config import settings


def _norm_key(v: Any) -> str:
    return str(v if v is not None else "").strip().upper()


# Annual-equivalent multipliers by WAGE_UNIT_OF_PAY (DOL uses these labels; tolerate abbreviations).
_WAGE_MULT = {"YEAR": 1, "YR": 1, "ANNUAL": 1, "HOUR": 2080, "HR": 2080, "WEEK": 52, "WK": 52,
              "BI-WEEKLY": 26, "BIWEEKLY": 26, "MONTH": 12, "MTH": 12, "MONTHLY": 12}


def _annual_wage(amount: Any, unit: Any) -> float | None:
    try:
        a = float(str(amount).replace(",", "").replace("$", "").strip())
    except (TypeError, ValueError):
        return None
    if a <= 0:
        return None
    mult = _WAGE_MULT.get(_norm_key(unit))
    if mult is None:                       # unit missing/unknown: guess (annual if big, else hourly)
        mult = 1 if a > 2000 else 2080
    val = a * mult
    return val if 1000 <= val <= 10_000_000 else None    # drop obvious garbage


def _is_certified(status: Any) -> bool:
    return _norm_key(status).startswith("CERTIFIED")     # "Certified" or "Certified - Withdrawn"


def _pick(row: dict, *names: str) -> Any:
    for n in names:
        v = row.get(n)
        if v not in (None, ""):
            return v
    return None


def _iter_rows(path: str) -> Iterator[dict]:
    """Yield each data row as {UPPER_HEADER: value}. Streams .csv (stdlib) or .xlsx (openpyxl)."""
    if path.lower().endswith(".csv"):
        import csv
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
            header: list[str] | None = None
            for row in csv.reader(fh):
                if header is None:
                    header = [_norm_key(c) for c in row]
                    continue
                yield dict(zip(header, row))
        return
    try:
        import openpyxl
    except ImportError as exc:                           # pragma: no cover - prod has openpyxl
        raise RuntimeError("openpyxl is required to read .xlsx files (pip install openpyxl), "
                           "or convert the file to .csv first.") from exc
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        header = None
        for row in wb.active.iter_rows(values_only=True):
            if header is None:
                header = [_norm_key(c) for c in row]
                continue
            yield dict(zip(header, row))
    finally:
        wb.close()


def _resolve(source: str) -> tuple[str, bool]:
    """Return (local_path, is_temp). Downloads http(s) sources to a temp file (streamed)."""
    if not source.lower().startswith(("http://", "https://")):
        return source, False
    suffix = ".csv" if source.lower().split("?")[0].endswith(".csv") else ".xlsx"
    fd, tmp = tempfile.mkstemp(suffix=suffix, prefix="h1b_")
    os.close(fd)
    with httpx.stream("GET", source, follow_redirects=True,
                      headers={"User-Agent": settings.scraper_user_agent},
                      timeout=settings.scraper_timeout_seconds) as r:
        r.raise_for_status()
        with open(tmp, "wb") as out:
            for chunk in r.iter_bytes(chunk_size=1 << 16):
                out.write(chunk)
    return tmp, True


def _aggregate(path: str, max_rows: int | None = None) -> dict[str, Any]:
    employers: Counter = Counter()
    occ_wages: dict[str, list[float]] = defaultdict(list)
    states: Counter = Counter()
    total = 0
    for row in _iter_rows(path):
        if _norm_key(_pick(row, "VISA_CLASS")) != "H-1B":          # exclude H-1B1 / E-3
            continue
        if not _is_certified(_pick(row, "CASE_STATUS")):
            continue
        total += 1
        emp = (str(_pick(row, "EMPLOYER_NAME") or "")).strip()
        if emp:
            employers[" ".join(emp.upper().split())] += 1
        wage = _annual_wage(_pick(row, "WAGE_RATE_OF_PAY_FROM", "WAGE_RATE_OF_PAY"),
                            _pick(row, "WAGE_UNIT_OF_PAY", "WAGE_UNIT"))
        occ = (str(_pick(row, "SOC_TITLE", "JOB_TITLE") or "")).strip()
        if occ and wage:
            occ_wages[" ".join(occ.title().split())].append(wage)
        st = _norm_key(_pick(row, "WORKSITE_STATE", "WORKSITE_STATE_1", "EMPLOYER_STATE"))
        if st:
            states[st] += 1
        if max_rows and total >= max_rows:
            break
    return {"total": total, "employers": employers, "occ_wages": occ_wages, "states": states}


def _to_knowledge(agg: dict, fiscal_year: str, top_n: int) -> int:
    from . import knowledge
    fy = f" (FY{fiscal_year})" if fiscal_year else ""
    docs = 0

    employers, total = agg["employers"], agg["total"]
    if employers:
        top = "; ".join(f"{name.title()} ({n:,})" for name, n in employers.most_common(top_n))
        if knowledge.upsert_document(
                source_type="dol_h1b", source_ref="h1b:employers", vertical=None,
                title="Top H-1B sponsoring employers in the USA",
                content=(f"Largest H-1B visa sponsors by certified labor condition applications{fy}, "
                         f"from U.S. Department of Labor public disclosure data — many are Indian-"
                         f"origin IT and consulting firms: {top}.")).get("ok"):
            docs += 1

    occ_wages = agg["occ_wages"]
    if occ_wages:
        ranked = sorted(occ_wages.items(), key=lambda kv: len(kv[1]), reverse=True)[:15]
        lines = []
        for occ, wages in ranked:
            med = statistics.median(wages)
            lines.append(f"{occ}: median offered wage about ${med:,.0f}/yr ({len(wages):,} filings)")
        all_w = [w for ws in occ_wages.values() for w in ws]
        overall = f"Across all occupations the median offered H-1B wage is about " \
                  f"${statistics.median(all_w):,.0f}/yr. " if all_w else ""
        if knowledge.upsert_document(
                source_type="dol_h1b", source_ref="h1b:wages", vertical=None,
                title="Typical H-1B wages by occupation",
                content=(f"Typical H-1B offered wages by occupation{fy}, from U.S. Department of "
                         f"Labor disclosure data. {overall}Most-common occupations: "
                         + "; ".join(lines) + ".")).get("ok"):
            docs += 1

    states = agg["states"]
    if states:
        top = "; ".join(f"{st} ({n:,})" for st, n in states.most_common(15))
        if knowledge.upsert_document(
                source_type="dol_h1b", source_ref="h1b:states", vertical=None,
                title="Where H-1B jobs are in the USA",
                content=(f"States with the most certified H-1B positions{fy} (U.S. Department of "
                         f"Labor disclosure data): {top}. "
                         + (f"About {total:,} H-1B positions were certified in total." if total else "")
                         )).get("ok"):
            docs += 1
    return docs


def import_disclosure(source: str | None = None, fiscal_year: str | None = None,
                      top_n: int = 25, max_rows: int | None = None) -> dict[str, Any]:
    """Download/parse the DOL H-1B disclosure file, aggregate, and feed Dost's knowledge base."""
    src = (source or settings.dol_h1b_disclosure_url or "").strip()
    if not src:
        return {"ok": False, "skipped": "no_source", "hint": "set DOL_H1B_DISCLOSURE_URL"}
    path, is_temp = _resolve(src)
    try:
        agg = _aggregate(path, max_rows=max_rows)
    finally:
        if is_temp:
            try:
                os.unlink(path)
            except OSError:
                pass
    fy = (fiscal_year or settings.dol_h1b_fiscal_year or "").strip()
    kb = _to_knowledge(agg, fy, top_n)
    return {"ok": True, "certified_h1b": agg["total"], "employers": len(agg["employers"]),
            "occupations": len(agg["occ_wages"]), "kb_documents": kb, "fiscal_year": fy or None}
